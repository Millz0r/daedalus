import daedalus.exceptions
import json
import os
import StringIO
import sys
import unittest

from daedalus.xlstransform import application, transforms, is_json
from mock import patch
from utils import xls_data, xlsx_data, json_xls_data, json_xlsx_data, to_json, json_corner_data, json_object_notation
import xlrd
import openpyxl

class TestXlstransform(unittest.TestCase):
    def test_xls_transform(self):
        transform = transforms.transform(xls_data(raw=True))
        self.assertTrue(isinstance(transform, basestring))
        self.assertEqual(to_json(transform), json_xls_data())

    def test_xlsx_transform(self):
        transform = transforms.transform(xlsx_data(raw=True))
        self.assertTrue(isinstance(transform, basestring))
        self.assertEqual(to_json(transform), json_xlsx_data())

    def test_json_transform_with_xlsx_data(self):
        with self.assertRaisesRegexp(daedalus.exceptions.BadFileFormat, 'Need JSON document to translate to excel'):
            transforms.transform(xlsx_data(raw=True), target_format='excel')

    def test_excel_transform_with_json_data(self):
        with self.assertRaisesRegexp(daedalus.exceptions.BadFileFormat, 'Need excel document to translate to JSON'):
            transforms.transform(json_xls_data(raw=True), source_format='json', target_format='json')

    def test_pretty_transform(self):
        transform = transforms.transform(xls_data(raw=True), pretty=True)
        self.assertTrue(isinstance(transform, basestring))
        self.assertEqual(to_json(transform), json_xls_data())

    def test_parse_args(self):
        args = application.parse_args([])
        self.assertTrue(isinstance(args.infile, file))
        self.assertEqual(args.infile.mode, 'r')
        self.assertTrue(isinstance(args.outfile, StringIO.StringIO))

    def test_main_json(self):
        with patch('daedalus.xlstransform.transform') as transform_mock:
            with patch('daedalus.xlstransform.application.parse_args') as parse_mock:
                with patch('daedalus.xlstransform.utils.is_json', return_value=True) as json_mock:
                    application.main()
                    parse_mock.assert_called()
                    json_mock.assert_called()
                    transform_mock.assert_called()

    def test_main_excel(self):
        with patch('daedalus.xlstransform.transform') as transform_mock:
            with patch('daedalus.xlstransform.application.parse_args') as parse_mock:
                with patch('daedalus.xlstransform.utils.is_json', return_value=False) as json_mock:
                    application.main()
                    parse_mock.assert_called()
                    json_mock.assert_called()
                    transform_mock.assert_called()

    def test_json_transform_typical_sample(self):
        data = transforms.transform(json_xls_data(raw=True), source_format='json', target_format='excel')
        self.assertIsInstance(data, basestring)
        self.assertIsInstance(openpyxl.load_workbook(StringIO.StringIO(data)), openpyxl.workbook.Workbook)

    def test_json_transform_corner_cases(self):
        # Test corner cases: blank cells, empty sheet, no rows - only columns
        data = transforms.transform(json_corner_data(raw=True), source_format='json', target_format='excel')
        self.assertIsInstance(data, basestring)
        self.assertIsInstance(openpyxl.load_workbook(StringIO.StringIO(data)), openpyxl.workbook.Workbook)

    def test_json_transform_object_notation(self):
        data = transforms.transform(json_object_notation(raw=True), source_format='json', target_format='excel')
        self.assertIsInstance(data, basestring)
        self.assertIsInstance(openpyxl.load_workbook(StringIO.StringIO(data)), openpyxl.workbook.Workbook)

    def test_is_json_true(self):
        self.assertTrue(is_json(json.dumps({'a': 'b'})))

    def test_is_json_false(self):
        self.assertFalse(is_json('oasfdasdf'))
